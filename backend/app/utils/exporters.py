"""
Export utilities — Excel (.xlsx) and PDF for BOQ and BBS.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


# ── Colour constants ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def export_boq_excel(boq: dict, project_name: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # Title
    ws.append([f"BILL OF QUANTITIES — {project_name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Section: {boq['section']}", "", f"Currency: {boq['currency']}"])
    ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"])
    ws.append([])

    headers = ["#", "Description", "Unit", "Quantity", "Unit Rate (ETB)", "Amount (ETB)", "Notes"]
    ws.append(headers)
    header_row = ws.max_row
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, line in enumerate(boq["lines"]):
        row = [
            line["item_number"],
            line["description"],
            line["unit"],
            line["quantity"],
            line["rate"],
            line["amount"],
            line.get("notes", ""),
        ]
        ws.append(row)
        if i % 2 == 0:
            for col_idx in range(1, len(row) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = ALT_FILL

    # Totals row
    ws.append([])
    ws.append(["", "TOTAL", "", "", "", boq["total_amount"], ""])
    total_row = ws.max_row
    for col_idx in range(1, 8):
        ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_bbs_excel(bars: list[dict], cutting_list: list[dict], project_name: str) -> bytes:
    wb = Workbook()

    # BBS sheet
    ws_bbs = wb.active
    ws_bbs.title = "BBS"
    ws_bbs.append([f"BAR BENDING SCHEDULE — {project_name}"])
    ws_bbs["A1"].font = Font(bold=True, size=14)
    ws_bbs.append([])

    bbs_headers = ["Bar Mark", "Member", "Dia (mm)", "Shape", "Qty",
                   "Clear Length (m)", "Cutting Length (m)", "Wt/Unit (kg)", "Total Wt (kg)", "Lap (mm)", "Notes"]
    ws_bbs.append(bbs_headers)
    hr = ws_bbs.max_row
    for ci in range(1, len(bbs_headers) + 1):
        c = ws_bbs.cell(row=hr, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    for bar in bars:
        ws_bbs.append([
            bar.get("bar_mark", ""),
            bar.get("member_name", ""),
            bar.get("bar_diameter_mm"),
            bar.get("bar_shape"),
            bar.get("quantity"),
            bar.get("clear_length_m"),
            bar.get("cutting_length_m"),
            bar.get("weight_per_unit_kg"),
            bar.get("total_weight_kg"),
            bar.get("lap_length_mm"),
            bar.get("notes", ""),
        ])

    _auto_width(ws_bbs)

    # Cutting list sheet
    ws_cl = wb.create_sheet("Cutting List")
    ws_cl.append([f"CUTTING LIST — {project_name}"])
    ws_cl["A1"].font = Font(bold=True, size=14)
    ws_cl.append([])
    cl_headers = ["Dia (mm)", "Cutting Length (m)", "Total Qty", "Total Weight (kg)"]
    ws_cl.append(cl_headers)
    for ci in range(1, len(cl_headers) + 1):
        c = ws_cl.cell(row=ws_cl.max_row, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    for item in cutting_list:
        ws_cl.append([item["diameter_mm"], item["cutting_length_m"], item["total_qty"], item["total_weight_kg"]])

    _auto_width(ws_cl)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_boq_pdf(boq: dict, project_name: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f"BOQ — {project_name}")
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>BILL OF QUANTITIES</b>", styles["Title"]))
    story.append(Paragraph(f"Project: {project_name} | Section: {boq['section']} | Currency: {boq['currency']}", styles["Normal"]))
    story.append(Spacer(1, 12))

    table_data = [["#", "Description", "Unit", "Qty", "Rate (ETB)", "Amount (ETB)"]]
    for line in boq["lines"]:
        table_data.append([
            str(line["item_number"]),
            line["description"],
            line["unit"],
            f"{line['quantity']:.3f}",
            f"{line['rate']:,.2f}",
            f"{line['amount']:,.2f}",
        ])
    table_data.append(["", "TOTAL", "", "", "", f"{boq['total_amount']:,.2f}"])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#D9E1F2")]),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()
